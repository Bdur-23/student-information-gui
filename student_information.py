import pandas as pd
import os
import tkinter as tk
from tkinter import messagebox
from openpyxl import load_workbook

def save_to_excel():
    name = name_entry.get()
    age = age_entry.get()
    phone = phone_entry.get()
    email = email_entry.get()

    data = {'Name': [name], 'Age': [age], 'Phone': [phone], 'Email': [email]}
    df = pd.DataFrame(data)

    file_exists = os.path.exists('student_info.xlsx')

    try:
        with pd.ExcelWriter('student_info.xlsx', engine='openpyxl', mode='a' if file_exists else 'w') as writer:
            if not file_exists:
                df.to_excel(writer, sheet_name='Sheet1', index=False, header=True)
            else:
                writer.book = load_workbook(writer.path)
                df.to_excel(writer, index=False, header=False, startrow=writer.sheets['Sheet1'].max_row + 1)

        messagebox.showinfo("Success", "Student information saved successfully!")

        # Clear input fields
        name_entry.delete(0, tk.END)
        age_entry.delete(0, tk.END)
        phone_entry.delete(0, tk.END)
        email_entry.delete(0, tk.END)

    except Exception as e:
        messagebox.showerror("Error", f"An error occurred: {str(e)}")

root = tk.Tk()
root.title("Student Information")
root.geometry("1920x1080")  # Full screen
root.attributes('-fullscreen', True)

frame = tk.Frame(root)
frame.pack(pady=50)

name_label = tk.Label(frame, text="Name:", font=("Helvetica", 16))
name_label.grid(row=0, column=0, padx=20)
name_entry = tk.Entry(frame, font=("Helvetica", 16))
name_entry.grid(row=0, column=1)

age_label = tk.Label(frame, text="Age:", font=("Helvetica", 16))
age_label.grid(row=1, column=0, padx=20)
age_entry = tk.Entry(frame, font=("Helvetica", 16))
age_entry.grid(row=1, column=1)

phone_label = tk.Label(frame, text="Phone:", font=("Helvetica", 16))
phone_label.grid(row=2, column=0, padx=20)
phone_entry = tk.Entry(frame, font=("Helvetica", 16))
phone_entry.grid(row=2, column=1)

email_label = tk.Label(frame, text="Email:", font=("Helvetica", 16))
email_label.grid(row=3, column=0, padx=20)
email_entry = tk.Entry(frame, font=("Helvetica", 16))
email_entry.grid(row=3, column=1)

save_button = tk.Button(frame, text="Save", font=("Helvetica", 16), command=save_to_excel)
save_button.grid(row=4, columnspan=2, pady=20)

exit_button = tk.Button(frame, text="Exit", font=("Helvetica", 16), command=root.destroy)
exit_button.grid(row=5, columnspan=2)

root.mainloop()
